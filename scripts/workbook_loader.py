"""
Workbook loader for metadata-driven CHA rendering.

Supports both:
- Normalized workbook metadata sheets (`_registry`, `_figure_specs`, `_table_specs`)
- Flat per-indicator sheets used by the Mid-Hudson CHA workbook template
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import json
import re

import pandas as pd

from scripts.cha_table_styling import CHA_REGION_ALIASES, CHA_REGION_ORDER


VALID_OBJECT_TYPES = {"table", "figure"}
VALID_FIGURE_TYPES = {
    "line",
    "clustered_bar",
    "stacked_bar",
    "simple_bar",
    "horizontal_bar",
    "horizontal_stacked_bar",
    "horizontal_clustered_bar",
    "pie",
    "dot_whisker",
}
_VALID_FORMAT_CODES: frozenset[str] = frozenset(
    {
        "integer",
        "number",
        "year",
        "percent1",
        "percent2",
        "currency",
        "currency2",
        "ratio",
        "date",
        "text",
    }
)

_X_COL_SYNONYM_GROUPS: tuple[frozenset[str], ...] = (
    frozenset({"year", "period"}),
    frozenset({"county", "region", "region/county"}),
    frozenset({"group", "category"}),
)


def _are_x_col_synonyms(left: str, right: str) -> bool:
    """True when two X Column / header labels refer to the same axis dimension."""
    a = _as_text(left).lower()
    b = _as_text(right).lower()
    if not a or not b:
        return False
    if a == b:
        return True
    return any(a in group and b in group for group in _X_COL_SYNONYM_GROUPS)


def _first_col_values_are_regions(series: pd.Series) -> bool:
    non_null = [_as_text(value) for value in series.tolist() if _as_text(value)]
    if not non_null:
        return False
    matches = sum(
        1
        for value in non_null
        if CHA_REGION_ALIASES.get(value, value) in CHA_REGION_ORDER
    )
    return matches / len(non_null) >= 0.8


_FIGURE_TYPE_ALIASES: dict[str, str] = {
    "1": "line",
    "line": "line",
    "2": "clustered_bar",
    "clustered_bar": "clustered_bar",
    "clustered bar": "clustered_bar",
    "cluster bar": "clustered_bar",
    "3": "stacked_bar",
    "stacked_bar": "stacked_bar",
    "stacked bar": "stacked_bar",
    "stack bar": "stacked_bar",
    "4": "simple_bar",
    "simple_bar": "simple_bar",
    "simple bar": "simple_bar",
    "bar": "simple_bar",
    "5": "horizontal_bar",
    "horizontal_bar": "horizontal_bar",
    "horizontal bar": "horizontal_bar",
    "6": "dot_whisker",
    "dot_whisker": "dot_whisker",
    "dot whisker": "dot_whisker",
    "dot plot": "dot_whisker",
    "dot_plot": "dot_whisker",
    "error bar": "dot_whisker",
    "error_bar": "dot_whisker",
    "7": "horizontal_stacked_bar",
    "horizontal_stacked_bar": "horizontal_stacked_bar",
    "horizontal stacked bar": "horizontal_stacked_bar",
    "horizontal stacked": "horizontal_stacked_bar",
    "stacked horizontal bar": "horizontal_stacked_bar",
    "8": "horizontal_clustered_bar",
    "horizontal_clustered_bar": "horizontal_clustered_bar",
    "horizontal clustered bar": "horizontal_clustered_bar",
    "clustered horizontal bar": "horizontal_clustered_bar",
    "cluster horizontal bar": "horizontal_clustered_bar",
    "horizontal cluster bar": "horizontal_clustered_bar",
    "9": "pie",
    "pie": "pie",
    "pie chart": "pie",
    "pie_chart": "pie",
}


@dataclass(frozen=True)
class RegistryRecord:
    object_id: str
    object_type: str
    label: str
    caption: str
    data_sheet: str
    enabled: bool
    section_tag: str
    order_index: int


@dataclass(frozen=True)
class FigureSpec:
    object_id: str
    figure_type: str
    x_col: str
    y_cols: list[str]
    x_axis_title: str
    y_axis_title: str
    start_at_zero: bool
    hover_suffix: str
    pivot_for_chart: bool = False
    show_data_labels: bool | None = None


@dataclass(frozen=True)
class CellStyle:
    """Excel cell formatting metadata for table rendering."""

    bold: bool = False
    indent: int = 0


@dataclass(frozen=True)
class CellMerge:
    """A merged cell region relative to a table grid origin (0-based)."""

    row: int
    col: int
    rowspan: int
    colspan: int


@dataclass(frozen=True)
class MergedTableGrid:
    """Rectangular cell grid with Excel merge metadata for live table rendering."""

    cells: tuple[tuple[Any, ...], ...]
    merges: tuple[CellMerge, ...]
    header_rows: int
    format_rules_by_col: tuple[str, ...] = ()
    styles: tuple[tuple[CellStyle, ...], ...] | None = None


@dataclass(frozen=True)
class TableSpec:
    object_id: str
    has_multilevel_headers: bool
    format_rules: dict[str, str]
    row_label_col: str
    merged_grid: MergedTableGrid | None = None
    cell_styles: tuple[tuple[CellStyle, ...], ...] | None = None


@dataclass(frozen=True)
class SourceSpec:
    object_id: str
    table_id: str
    url: str
    data_year: int
    estimate_type: str
    citation_month: str
    citation_year: int
    custom_text: str
    source_text: str = ""
    note_text: str = ""


@dataclass(frozen=True)
class WorkbookModel:
    workbook_path: Path
    registry: dict[str, RegistryRecord]
    figure_specs: dict[str, FigureSpec]
    table_specs: dict[str, TableSpec]
    source_specs: dict[str, SourceSpec]
    data_frames: dict[str, pd.DataFrame]


def _as_bool(value: Any, default: bool = False) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text == "":
        return default
    return text in {"1", "true", "yes", "y", "enabled"}


def _as_text(value: Any, default: str = "") -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    # Whole-number floats (e.g. year headers 2020.0) render as plain integers.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_int(value: Any, default: int = 0) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    text = str(value).strip()
    if text == "":
        return default
    try:
        return int(float(text.replace(",", "")))
    except ValueError:
        # Accept year ranges like "2021-2023" / "2021–2023" by taking the first year.
        match = re.search(r"\d{4}", text)
        if match:
            return int(match.group(0))
        return default


def _parse_string_list(value: Any) -> list[str]:
    text = _as_text(value)
    if not text:
        return []
    return [item.strip() for item in text.split(",") if item.strip()]


def _normalize_figure_type(value: Any, default: str | None = "line") -> str:
    text = _as_text(value).strip().lower()
    if not text:
        return default or ""
    normalized = _FIGURE_TYPE_ALIASES.get(text)
    if normalized:
        return normalized
    text = text.replace("-", "_")
    if text in VALID_FIGURE_TYPES:
        return text
    return default or ""


def _group_by_to_pivot_for_chart(group_by_value: Any) -> bool | None:
    """
    Map optional grouping intent metadata to pivot_for_chart semantics.

    - group_by = x_col  -> group bars by x_col values (pivot required)
    - group_by = series -> group bars by existing series columns (no pivot)
    """
    text = _as_text(group_by_value).strip().lower()
    if not text:
        return None
    normalized = text.replace(" ", "_").replace("-", "_")
    if normalized == "x_col":
        return True
    if normalized == "series":
        return False
    return None


def _flat_slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def _build_master_indicator_lookup(
    master_df: pd.DataFrame | None,
) -> dict[str, tuple[str, str]]:
    """Map indicator Name -> (Object ID, Figure ID) from the Master sheet."""
    if master_df is None or master_df.empty:
        return {}

    lookup: dict[str, tuple[str, str]] = {}
    for _, row in master_df.iterrows():
        name = _as_text(row.iloc[5] if master_df.shape[1] > 5 else None)
        object_id = _as_text(row.iloc[6] if master_df.shape[1] > 6 else None)
        figure_id = _as_text(row.iloc[7] if master_df.shape[1] > 7 else None)
        if name:
            lookup[name] = (object_id, figure_id)
    return lookup


def _normalize_config_key(text: Any) -> str:
    raw = _as_text(text).lower()
    return re.sub(r"[^a-z0-9]+", " ", raw).strip()


def _config_value(config: dict[str, Any], key: str, default: Any = None) -> Any:
    if key in config:
        return config[key]
    normalized_key = _normalize_config_key(key)
    for existing_key, value in config.items():
        if _normalize_config_key(existing_key) == normalized_key:
            return value
    return default


_FLAT_DATA_COL_START = 5  # column F (0-based index 5)
_FLAT_ENTER_DATA_MARKER = "enter data"


def _flat_data_bounds(df: pd.DataFrame) -> tuple[int, int, int] | None:
    """Return (header_row_idx, data_col_start, data_col_end) for a flat sheet."""
    data_col_start_idx = _FLAT_DATA_COL_START
    data_header_idx: int | None = None
    for i in range(len(df)):
        if df.shape[1] > 4 and _as_text(df.iloc[i, 4]).lower() == _FLAT_ENTER_DATA_MARKER:
            data_header_idx = i
            break
    if data_header_idx is None or df.shape[1] <= data_col_start_idx:
        return None

    header_scan_rows = range(data_header_idx, min(len(df), data_header_idx + 6))
    last_used_col = data_col_start_idx - 1
    for j in range(data_col_start_idx, df.shape[1]):
        if any(_as_text(df.iloc[i, j]) != "" for i in header_scan_rows):
            last_used_col = j
    if last_used_col < data_col_start_idx:
        return None
    return data_header_idx, data_col_start_idx, last_used_col + 1


def _flat_last_header_row_idx(df: pd.DataFrame, header_row_idx: int, data_col_end_idx: int) -> int:
    """Best-effort last header row before data rows (supports two-row headers)."""
    data_col_start_idx = _FLAT_DATA_COL_START
    sub_header_row_idx = header_row_idx + 1
    if sub_header_row_idx >= len(df):
        return header_row_idx

    row_label_blank = _as_text(df.iloc[sub_header_row_idx, data_col_start_idx]) == ""
    has_data_headers = any(
        _as_text(df.iloc[sub_header_row_idx, j])
        for j in range(data_col_start_idx + 1, data_col_end_idx)
    )
    if row_label_blank and has_data_headers:
        return sub_header_row_idx
    return header_row_idx


def _worksheet_has_data_body_merges(
    ws: Any,
    *,
    header_row_1based: int,
    last_header_row_1based: int,
    data_col_start_1based: int,
) -> bool:
    """True when merges extend below the header block into table body rows."""
    for mc in ws.merged_cells.ranges:
        if mc.min_col < data_col_start_1based:
            continue
        if mc.min_row > last_header_row_1based:
            return True
    return False


def _resolved_worksheet_value(ws: Any, row_1based: int, col_1based: int) -> Any:
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row_1based <= mc.max_row and mc.min_col <= col_1based <= mc.max_col:
            return ws.cell(mc.min_row, mc.min_col).value
    return ws.cell(row_1based, col_1based).value


def _style_anchor_cell(ws: Any, row_1based: int, col_1based: int) -> tuple[int, int]:
    """Return the top-left anchor of a merged region, or the cell itself."""
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= row_1based <= mc.max_row and mc.min_col <= col_1based <= mc.max_col:
            return mc.min_row, mc.min_col
    return row_1based, col_1based


def _read_cell_style(ws: Any, row_1based: int, col_1based: int) -> CellStyle:
    anchor_row, anchor_col = _style_anchor_cell(ws, row_1based, col_1based)
    cell = ws.cell(anchor_row, anchor_col)
    bold = bool(cell.font and cell.font.bold)
    indent = 0
    if cell.alignment and cell.alignment.indent:
        indent = int(cell.alignment.indent)
    return CellStyle(bold=bold, indent=indent)


def _extract_data_block_styles(
    ws: Any,
    *,
    data_start_idx: int,
    data_col_start_idx: int,
    n_rows: int,
    n_cols: int,
) -> tuple[tuple[CellStyle, ...], ...]:
    """Build a style grid aligned with parsed data_df rows and columns."""
    styles: list[tuple[CellStyle, ...]] = []
    for row_offset in range(n_rows):
        row_1based = data_start_idx + row_offset + 1
        row_styles: list[CellStyle] = []
        for col_offset in range(n_cols):
            col_1based = data_col_start_idx + col_offset + 1
            row_styles.append(_read_cell_style(ws, row_1based, col_1based))
        styles.append(tuple(row_styles))
    return tuple(styles)


def _flat_grid_stop_row(ws: Any, start_row_1based: int, data_col_start_1based: int, data_col_end_1based: int) -> int:
    """Find the last populated data row in the flat-sheet data block.

    Flat indicator sheets interleave config keys in column A with table data in
    columns F+, so only blank streaks in the data columns end the block.
    """
    last_row = start_row_1based
    blank_streak = 0
    for row_1based in range(start_row_1based + 1, ws.max_row + 1):
        row_vals = [
            _resolved_worksheet_value(ws, row_1based, col_1based)
            for col_1based in range(data_col_start_1based, data_col_end_1based)
        ]
        if all(_as_text(value) == "" for value in row_vals):
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        last_row = row_1based
    return last_row


def _extend_data_col_end_for_header_merges(
    ws: Any,
    *,
    header_row_1based: int,
    data_col_start_1based: int,
    data_col_end_1based: int,
) -> int:
    """Include trailing merge partners that can be blank in the header row."""
    extended = data_col_end_1based
    for mc in ws.merged_cells.ranges:
        if mc.min_row <= header_row_1based <= mc.max_row and mc.min_col >= data_col_start_1based:
            extended = max(extended, mc.max_col + 1)
    return extended


def _extract_merged_table_grid(ws: Any, df: pd.DataFrame) -> MergedTableGrid | None:
    """Build a merged cell grid when Excel merges extend into table body rows."""
    bounds = _flat_data_bounds(df)
    if bounds is None:
        return None

    header_row_idx, data_col_start_idx, data_col_end_idx = bounds
    header_row_1based = header_row_idx + 1
    data_col_start_1based = data_col_start_idx + 1
    data_col_end_1based = _extend_data_col_end_for_header_merges(
        ws,
        header_row_1based=header_row_1based,
        data_col_start_1based=data_col_start_1based,
        data_col_end_1based=data_col_end_idx + 1,
    )
    last_header_row_1based = _flat_last_header_row_idx(df, header_row_idx, data_col_end_idx) + 1

    if not _worksheet_has_data_body_merges(
        ws,
        header_row_1based=header_row_1based,
        last_header_row_1based=last_header_row_1based,
        data_col_start_1based=data_col_start_1based,
    ):
        return None

    last_row_1based = _flat_grid_stop_row(ws, header_row_1based, data_col_start_1based, data_col_end_1based)
    n_rows = last_row_1based - header_row_1based + 1
    n_cols = data_col_end_1based - data_col_start_1based
    if n_rows <= 0 or n_cols <= 0:
        return None

    grid: list[list[Any]] = []
    style_grid: list[list[CellStyle]] = []
    for row_1based in range(header_row_1based, last_row_1based + 1):
        grid.append(
            [
                _resolved_worksheet_value(ws, row_1based, col_1based)
                for col_1based in range(data_col_start_1based, data_col_end_1based)
            ]
        )
        style_grid.append(
            [
                _read_cell_style(ws, row_1based, col_1based)
                for col_1based in range(data_col_start_1based, data_col_end_1based)
            ]
        )

    merges: list[CellMerge] = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row < header_row_1based or mc.max_row > last_row_1based:
            continue
        if mc.min_col < data_col_start_1based or mc.max_col >= data_col_end_1based:
            continue
        merges.append(
            CellMerge(
                row=mc.min_row - header_row_1based,
                col=mc.min_col - data_col_start_1based,
                rowspan=mc.max_row - mc.min_row + 1,
                colspan=mc.max_col - mc.min_col + 1,
            )
        )

    format_row_idx = header_row_idx - 1
    format_rules_by_col: list[str] = []
    if format_row_idx >= 0:
        for col_1based in range(data_col_start_1based, data_col_end_1based):
            fmt_val = _as_text(ws.cell(format_row_idx + 1, col_1based).value).lower()
            format_rules_by_col.append(fmt_val if fmt_val in _VALID_FORMAT_CODES else "")

    # Include detected sub-header rows (e.g. None/One/Two under a merged group).
    header_rows = last_header_row_1based - header_row_1based + 1
    return MergedTableGrid(
        cells=tuple(tuple(row) for row in grid),
        merges=tuple(merges),
        header_rows=header_rows,
        format_rules_by_col=tuple(format_rules_by_col),
        styles=tuple(tuple(row) for row in style_grid),
    )


def _read_excel_raw(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return pd.read_excel(path, sheet_name=None, header=None, keep_default_na=False, na_values=[""])


def _read_excel_with_headers(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return pd.read_excel(path, sheet_name=None, keep_default_na=False, na_values=[""])


def _is_flat_indicator_sheet(df: pd.DataFrame) -> bool:
    return bool(df is not None and not df.empty and df.shape[1] >= 2 and _as_text(df.iloc[0, 0]).lower() == "name")


def _parse_flat_indicator_sheet(
    sheet_name: str, df: pd.DataFrame
) -> tuple[dict[str, Any], pd.DataFrame, dict[str, str], bool, tuple[int, int, int] | None]:
    config: dict[str, Any] = {}
    for i in range(len(df)):
        key = _as_text(df.iloc[i, 0] if df.shape[1] > 0 else None)
        if key:
            config[key] = df.iloc[i, 1] if df.shape[1] > 1 else None

    data_header_idx: int | None = None
    for i in range(len(df)):
        if df.shape[1] > 4 and _as_text(df.iloc[i, 4]) == "Enter Data":
            data_header_idx = i
            break

    # In flat template sheets, all columns to the right of "Enter Data"
    # are indicator data columns (first data column starts at index 5 / col F).
    data_col_start_idx = 5
    if data_header_idx is None or df.shape[1] <= data_col_start_idx:
        return config, pd.DataFrame(), {}, False, None

    # Determine right-most used data column from a small header/data window.
    # This avoids trailing blank worksheet columns becoming ghost headers.
    header_scan_rows = range(data_header_idx, min(len(df), data_header_idx + 6))
    last_used_col = data_col_start_idx - 1
    for j in range(data_col_start_idx, df.shape[1]):
        if any(_as_text(df.iloc[i, j]) != "" for i in header_scan_rows):
            last_used_col = j

    if last_used_col < data_col_start_idx:
        return config, pd.DataFrame(), {}, False, None

    data_col_end_idx = last_used_col + 1
    header_row_idx = data_header_idx
    format_row_idx = data_header_idx - 1
    raw_headers = [df.iloc[header_row_idx, j] for j in range(data_col_start_idx, data_col_end_idx)]
    headers = [_as_text(h) for h in raw_headers]

    # Some sheets place format codes (integer/percent1/etc.) on the same row as
    # "Enter Data" and put real headers on the next row.
    nonempty_headers = [h for h in headers if h]
    looks_like_format_header_row = bool(nonempty_headers) and all(
        _as_text(h).lower() in _VALID_FORMAT_CODES for h in nonempty_headers
    )
    if looks_like_format_header_row and header_row_idx + 1 < len(df):
        next_headers = [
            _as_text(df.iloc[header_row_idx + 1, j])
            for j in range(data_col_start_idx, data_col_end_idx)
        ]
        next_nonempty = [h for h in next_headers if h]
        next_is_real_headers = bool(next_nonempty) and not all(
            h.lower() in _VALID_FORMAT_CODES for h in next_nonempty
        )
        if next_is_real_headers:
            format_row_idx = header_row_idx
            header_row_idx = header_row_idx + 1
            headers = next_headers

    # Detect two-row merged header pattern (multilevel table).
    data_start_idx = header_row_idx + 1
    auto_multilevel = False
    sub_header_row_idx = data_header_idx + 1
    explicit_multilevel = (
        _as_bool(_config_value(config, "Multilevel Headers", False))
        or _as_bool(_config_value(config, "Multiheader Levels", False))
    )
    if sub_header_row_idx < len(df):
        row_label_blank = _as_text(df.iloc[sub_header_row_idx, data_col_start_idx]) == ""
        has_data_headers = any(
            _as_text(df.iloc[sub_header_row_idx, j])
            for j in range(data_col_start_idx + 1, data_col_end_idx)
        )

        def _looks_numeric_cell(value: Any) -> bool:
            if value is None or (isinstance(value, float) and pd.isna(value)):
                return False
            if isinstance(value, (int, float)):
                return True
            text = _as_text(value)
            if not text:
                return False
            text = text.replace(",", "").replace("$", "").rstrip("%").strip()
            text = re.sub(r"[*†‡§#]+$", "", text)
            try:
                float(text)
                return True
            except ValueError:
                return False

        # Primary detection: traditional template shape (blank first subheader cell).
        use_subheader_row = row_label_blank and has_data_headers

        # Secondary detection (enabled for explicit multilevel metadata):
        # some sheets repeat "County" in the first subheader cell.
        # Treat that row as header only when the following row looks numeric.
        if not use_subheader_row and explicit_multilevel and has_data_headers:
            candidate_cells = [
                df.iloc[sub_header_row_idx, j]
                for j in range(data_col_start_idx + 1, data_col_end_idx)
            ]
            candidate_nonempty = [cell for cell in candidate_cells if _as_text(cell)]
            candidate_numeric_count = sum(1 for cell in candidate_cells if _looks_numeric_cell(cell))

            next_row_idx = sub_header_row_idx + 1
            next_row_numeric_count = 0
            if next_row_idx < len(df):
                next_row_cells = [
                    df.iloc[next_row_idx, j]
                    for j in range(data_col_start_idx + 1, data_col_end_idx)
                ]
                next_row_numeric_count = sum(1 for cell in next_row_cells if _looks_numeric_cell(cell))

            has_label_like_subheaders = len(candidate_nonempty) >= 2 and candidate_numeric_count <= 1
            has_numeric_data_after = next_row_numeric_count >= 2
            use_subheader_row = has_label_like_subheaders and has_numeric_data_after

        if use_subheader_row:
            sub_headers = [
                _as_text(df.iloc[sub_header_row_idx, j])
                for j in range(data_col_start_idx, data_col_end_idx)
            ]
            merged: list[str] = []
            last_top = ""
            last_sub = ""
            for col_idx, (h_top, h_sub) in enumerate(zip(headers, sub_headers)):
                # New non-blank top group resets mid-level forward-fill so
                # categories do not leak across unrelated column groups.
                if h_top:
                    last_top = h_top
                    last_sub = h_sub
                elif h_sub:
                    last_sub = h_sub
                effective_top = last_top
                effective_sub = h_sub or last_sub
                if col_idx == 0:
                    merged.append(h_sub or h_top)
                elif effective_top and effective_sub:
                    merged.append(f"{effective_top}|{effective_sub}")
                elif effective_sub:
                    merged.append(effective_sub)
                else:
                    merged.append(effective_top)
            headers = merged
            data_start_idx = sub_header_row_idx + 1
            # Only treat as true multilevel when the top header is semantic
            # (e.g., "Three-Year Average"), not a format token (e.g., "percent1").
            auto_multilevel = any(
                "|" in col and _as_text(col.split("|", 1)[0]).lower() not in _VALID_FORMAT_CODES
                for col in headers
            )

        # Tertiary header row: category names on row 2, then "#" / "%" units on
        # row 3 (e.g. Educ Person 25+). Consume that unit row before data.
        unit_row_idx = data_start_idx
        if unit_row_idx < len(df):
            unit_row_label_blank = _as_text(df.iloc[unit_row_idx, data_col_start_idx]) == ""
            unit_cells = [
                df.iloc[unit_row_idx, j]
                for j in range(data_col_start_idx + 1, data_col_end_idx)
            ]
            unit_labels = [_as_text(cell) for cell in unit_cells if _as_text(cell)]
            unit_numeric_count = sum(1 for cell in unit_cells if _looks_numeric_cell(cell))
            known_unit_tokens = {"#", "%", "n", "no.", "number", "count", "rate"}
            looks_like_unit_row = (
                unit_row_label_blank
                and len(unit_labels) >= 2
                and unit_numeric_count == 0
                and all(
                    label.lower() in known_unit_tokens or len(label) <= 3
                    for label in unit_labels
                )
            )
            next_row_idx = unit_row_idx + 1
            next_row_numeric_count = 0
            if next_row_idx < len(df):
                next_row_cells = [
                    df.iloc[next_row_idx, j]
                    for j in range(data_col_start_idx + 1, data_col_end_idx)
                ]
                next_row_numeric_count = sum(
                    1 for cell in next_row_cells if _looks_numeric_cell(cell)
                )
            if looks_like_unit_row and next_row_numeric_count >= 2:
                unit_headers = [
                    _as_text(df.iloc[unit_row_idx, j])
                    for j in range(data_col_start_idx, data_col_end_idx)
                ]
                with_units: list[str] = []
                for col_idx, (h_base, h_unit) in enumerate(zip(headers, unit_headers)):
                    if col_idx == 0:
                        with_units.append(h_base)
                    elif h_base and h_unit:
                        with_units.append(f"{h_base}|{h_unit}")
                    elif h_unit:
                        with_units.append(h_unit)
                    else:
                        with_units.append(h_base)
                headers = with_units
                data_start_idx = unit_row_idx + 1
                auto_multilevel = True

    # Read per-column format rules from row above Enter Data.
    format_rules: dict[str, str] = {}
    if format_row_idx >= 0:
        fmt_row_idx = format_row_idx
        for j in range(data_col_start_idx, data_col_end_idx):
            header_pos = j - data_col_start_idx
            if header_pos < len(headers):
                fmt_val = _as_text(df.iloc[fmt_row_idx, j]).lower()
                if fmt_val in _VALID_FORMAT_CODES:
                    format_rules[headers[header_pos]] = fmt_val

    # Also infer format rules from tokenized headers: "percent1|Dutchess"
    for header in headers:
        text = _as_text(header)
        if "|" in text:
            left, _right = text.split("|", 1)
            fmt_name = _as_text(left).lower()
            if fmt_name in _VALID_FORMAT_CODES:
                format_rules[text] = fmt_name

    def _clean_header_label(label: Any) -> str:
        text = _as_text(label)
        if "|" in text:
            left, right = text.split("|", 1)
            if _as_text(left).lower() in _VALID_FORMAT_CODES:
                return _as_text(right)
        return text

    data_rows: list[list[Any]] = []
    for i in range(data_start_idx, len(df)):
        row_label = df.iloc[i, data_col_start_idx]
        if row_label is None or (isinstance(row_label, float) and pd.isna(row_label)) or _as_text(row_label) == "":
            break
        row_values = [
            df.iloc[i, j] if j < df.shape[1] else None
            for j in range(data_col_start_idx, data_col_start_idx + len(headers))
        ]
        data_rows.append(row_values)

    data_df = pd.DataFrame(data_rows, columns=headers) if data_rows else pd.DataFrame(columns=headers)
    cleaned_headers = [_clean_header_label(col) for col in data_df.columns]
    data_df = data_df.copy()
    data_df.columns = cleaned_headers

    remapped_rules: dict[str, str] = {}
    for col_name, fmt in format_rules.items():
        remapped_rules[_clean_header_label(col_name)] = fmt

    data_region = (data_start_idx, data_col_start_idx, data_col_end_idx)
    return config, data_df, remapped_rules, auto_multilevel, data_region


def _load_flat_workbook(source_path: Path) -> WorkbookModel:
    import openpyxl

    raw_sheets = _read_excel_raw(source_path)
    openpyxl_wb = openpyxl.load_workbook(source_path, data_only=True)
    skip_sheets = {"Master", "Dropdowns", "Template", "Template2", "_Template", "_Template (2)"}

    registry: dict[str, RegistryRecord] = {}
    figure_specs: dict[str, FigureSpec] = {}
    table_specs: dict[str, TableSpec] = {}
    source_specs: dict[str, SourceSpec] = {}
    data_frames: dict[str, pd.DataFrame] = {}
    order_counter = 0
    master_lookup = _build_master_indicator_lookup(raw_sheets.get("Master"))

    for sheet_name, sheet_df in raw_sheets.items():
        if sheet_name in skip_sheets:
            continue
        if sheet_name.startswith("_"):
            continue
        if not _is_flat_indicator_sheet(sheet_df):
            continue

        config, data_df, format_rules, auto_multilevel, data_region = _parse_flat_indicator_sheet(
            sheet_name, sheet_df
        )
        merged_grid: MergedTableGrid | None = None
        cell_styles: tuple[tuple[CellStyle, ...], ...] | None = None
        if sheet_name in openpyxl_wb.sheetnames:
            ws = openpyxl_wb[sheet_name]
            merged_grid = _extract_merged_table_grid(ws, sheet_df)
            if data_region and not data_df.empty:
                data_start_idx, data_col_start_idx, _data_col_end_idx = data_region
                cell_styles = _extract_data_block_styles(
                    ws,
                    data_start_idx=data_start_idx,
                    data_col_start_idx=data_col_start_idx,
                    n_rows=len(data_df),
                    n_cols=len(data_df.columns),
                )

        # When the user sets "X Column" but it doesn't match any data column,
        # reconcile the mismatch so the spec, y_cols, and renderer all agree.
        configured_x = _as_text(_config_value(config, "X Column", ""))
        # A comma in the configured value means it is a list of category
        # values, not a column name — skip the rename in that case.
        is_column_name = configured_x and "," not in configured_x
        pivot_for_chart = _as_bool(_config_value(config, "Pivot For Chart", False))
        if is_column_name and not data_df.empty and configured_x not in data_df.columns:
            first_col_name = data_df.columns[0]
            if first_col_name == "":
                new_cols = list(data_df.columns)
                # X Column names the post-pivot x-axis field. When regions are
                # stored as rows and pivot_for_chart is on, keep a neutral row
                # label column for the renderer to transpose.
                if pivot_for_chart and _first_col_values_are_regions(data_df.iloc[:, 0]):
                    new_cols[0] = "County"
                else:
                    new_cols[0] = configured_x
                data_df.columns = new_cols
            elif _are_x_col_synonyms(first_col_name, configured_x):
                # True synonym mismatch only (e.g., header "Year" vs X Column
                # "Period"). Do not rename across dimensions (County vs
                # Race/Ethnicity) — X Column names the intended x-axis field
                # after any pivot_for_chart reshape in the renderer.
                new_cols = list(data_df.columns)
                new_cols[0] = configured_x
                data_df.columns = new_cols

        raw_type = _as_text(_config_value(config, "Table/Figure/Both", "both")).lower()
        sheet_slug = _flat_slug(sheet_name)

        indicator_name = _as_text(_config_value(config, "Name", sheet_name))
        object_id_override = _as_text(_config_value(config, "Object ID", ""))
        figure_id_override = _as_text(_config_value(config, "Figure ID", "")) or object_id_override
        master_ids = master_lookup.get(indicator_name, ("", ""))
        if not object_id_override and master_ids[0]:
            object_id_override = master_ids[0]
        if not figure_id_override and master_ids[1]:
            figure_id_override = master_ids[1]
        if not figure_id_override:
            figure_id_override = object_id_override

        def _make_base_id(override: str) -> str:
            slug = _flat_slug(override) if override else sheet_slug
            for prefix in ("tbl-", "fig-"):
                if slug.startswith(prefix):
                    slug = slug[len(prefix):]
            return slug

        tbl_base_id = _make_base_id(object_id_override)
        fig_base_id = _make_base_id(figure_id_override)

        caption = _as_text(_config_value(config, "Name", sheet_name))
        section_tag = tbl_base_id
        data_sheet_name = f"data_{tbl_base_id}"
        data_frames[data_sheet_name] = data_df.copy()

        to_create: list[str] = []
        if raw_type in {"figure", "both"}:
            to_create.append("figure")
        if raw_type in {"table", "both"}:
            to_create.append("table")

        # Some flat-workbook sheets keep "Figure" in Table/Figure/Both while
        # still providing explicit tbl-/fig- IDs. Respect explicit IDs so both
        # include objects are generated when requested in metadata.
        object_id_lower = object_id_override.lower()
        figure_id_lower = figure_id_override.lower()
        if object_id_lower.startswith("tbl-") and "table" not in to_create:
            to_create.append("table")
        if figure_id_lower.startswith("fig-") and "figure" not in to_create:
            to_create.append("figure")

        for obj_type in to_create:
            base = fig_base_id if obj_type == "figure" else tbl_base_id
            prefix = "fig" if obj_type == "figure" else "tbl"
            object_id = f"{prefix}-{base}"
            registry[object_id] = RegistryRecord(
                object_id=object_id,
                object_type=obj_type,
                label=object_id,
                caption=caption,
                data_sheet=data_sheet_name,
                enabled=True,
                section_tag=section_tag,
                order_index=order_counter,
            )

            if obj_type == "figure":
                figure_type = _normalize_figure_type(_config_value(config, "Figure Type", "line"), default="line")
                x_col = _as_text(_config_value(config, "X Column", "")) or (data_df.columns[0] if not data_df.empty else "")
                y_cols_cfg = _parse_string_list(_config_value(config, "Y Column", ""))
                if y_cols_cfg:
                    y_cols = [col for col in y_cols_cfg if col in data_df.columns and col != x_col]
                else:
                    y_cols = [col for col in data_df.columns if col != x_col]
                group_by_override = _group_by_to_pivot_for_chart(_config_value(config, "Group By", ""))
                pivot_for_chart = (
                    group_by_override
                    if group_by_override is not None
                    else _as_bool(_config_value(config, "Pivot For Chart", False))
                )
                show_data_labels_raw = _config_value(config, "Show Data Labels", "")
                figure_specs[object_id] = FigureSpec(
                    object_id=object_id,
                    figure_type=figure_type,
                    x_col=x_col,
                    y_cols=y_cols,
                    x_axis_title=_as_text(_config_value(config, "X Axis Title", x_col)),
                    y_axis_title=_as_text(_config_value(config, "Y Axis Title", "")),
                    start_at_zero=_as_bool(_config_value(config, "Start at Zero", False)),
                    hover_suffix=_as_text(_config_value(config, "Hover Suffix", "%")),
                    pivot_for_chart=pivot_for_chart,
                    show_data_labels=(
                        _as_bool(show_data_labels_raw, default=False)
                        if _as_text(show_data_labels_raw) != ""
                        else None
                    ),
                )
            else:
                has_multilevel = auto_multilevel or _as_bool(_config_value(config, "Multilevel Headers", False))
                if merged_grid is not None:
                    has_multilevel = False
                # fallback by generic Data Type if no explicit format rules detected
                if not format_rules and not data_df.empty:
                    data_type = _as_text(_config_value(config, "Data Type", "")).lower()
                    data_cols = list(data_df.columns[1:])
                    if data_type in {"percent", "percentage"}:
                        format_rules = {col: "percent1" for col in data_cols}
                    elif data_type == "number":
                        format_rules = {col: "number" for col in data_cols}
                    elif data_type == "currency":
                        format_rules = {col: "currency" for col in data_cols}
                table_specs[object_id] = TableSpec(
                    object_id=object_id,
                    has_multilevel_headers=has_multilevel,
                    format_rules=format_rules,
                    row_label_col=(data_df.columns[0] if not data_df.empty else ""),
                    merged_grid=merged_grid,
                    cell_styles=cell_styles,
                )

            raw_source = _as_text(_config_value(config, "Source", ""))
            if raw_source.lower().startswith("source:"):
                raw_source = raw_source[len("source:"):].strip()

            raw_note = _as_text(_config_value(config, "Note", ""))

            source_specs[object_id] = SourceSpec(
                object_id=object_id,
                table_id=_as_text(_config_value(config, "Table ID", "")),
                url=_as_text(_config_value(config, "URL", "")),
                data_year=_as_int(_config_value(config, "Data Year", 2023), default=2023),
                estimate_type=_as_text(_config_value(config, "Estimate Type", "5-Year Estimates")),
                citation_month=_as_text(_config_value(config, "Citation Month", "April")),
                citation_year=_as_int(_config_value(config, "Citation Year", 2025), default=2025),
                custom_text=_as_text(_config_value(config, "Custom Text", "")),
                source_text=raw_source,
                note_text=raw_note,
            )

        order_counter += 1

    return WorkbookModel(
        workbook_path=source_path,
        registry=registry,
        figure_specs=figure_specs,
        table_specs=table_specs,
        source_specs=source_specs,
        data_frames=data_frames,
    )


def _load_normalized_workbook(source_path: Path) -> WorkbookModel:
    sheets = _read_excel_with_headers(source_path)
    required = {"_registry", "_figure_specs", "_table_specs"}
    if not required.issubset(sheets.keys()):
        missing = sorted(required - set(sheets.keys()))
        raise ValueError("Missing required metadata sheets: " + ", ".join(missing))

    registry_df = sheets["_registry"].copy()
    figure_df = sheets["_figure_specs"].copy()
    table_df = sheets["_table_specs"].copy()
    source_df = sheets.get("_source_specs", pd.DataFrame())

    registry: dict[str, RegistryRecord] = {}
    for row in registry_df.to_dict("records"):
        if not _as_bool(row.get("enabled", True), default=True):
            continue
        object_id = _as_text(row.get("object_id"))
        if not object_id:
            continue
        object_type = _as_text(row.get("object_type")).lower()
        if object_type not in VALID_OBJECT_TYPES:
            continue
        registry[object_id] = RegistryRecord(
            object_id=object_id,
            object_type=object_type,
            label=_as_text(row.get("label", object_id)),
            caption=_as_text(row.get("caption", "")),
            data_sheet=_as_text(row.get("data_sheet", "")),
            enabled=True,
            section_tag=_as_text(row.get("section_tag", "")),
            order_index=_as_int(row.get("order_index", 0), default=0),
        )

    figure_specs: dict[str, FigureSpec] = {}
    for row in figure_df.to_dict("records"):
        object_id = _as_text(row.get("object_id", ""))
        if object_id not in registry:
            continue
        group_by_override = _group_by_to_pivot_for_chart(row.get("group_by", ""))
        pivot_for_chart = (
            group_by_override
            if group_by_override is not None
            else _as_bool(row.get("pivot_for_chart", False))
        )
        figure_specs[object_id] = FigureSpec(
            object_id=object_id,
            figure_type=_normalize_figure_type(row.get("figure_type", "line"), default="line"),
            x_col=_as_text(row.get("x_col", "")),
            y_cols=_parse_string_list(row.get("y_cols", "")),
            x_axis_title=_as_text(row.get("x_axis_title", "")),
            y_axis_title=_as_text(row.get("y_axis_title", "")),
            start_at_zero=_as_bool(row.get("start_at_zero", False)),
            hover_suffix=_as_text(row.get("hover_suffix", "")),
            pivot_for_chart=pivot_for_chart,
            show_data_labels=(
                _as_bool(row.get("show_data_labels", False))
                if str(row.get("show_data_labels", "")).strip() != ""
                else None
            ),
        )

    table_specs: dict[str, TableSpec] = {}
    for row in table_df.to_dict("records"):
        object_id = _as_text(row.get("object_id", ""))
        if object_id not in registry:
            continue
        rules_raw = _as_text(row.get("format_rules_json", "{}"), default="{}")
        try:
            parsed_rules = json.loads(rules_raw) if rules_raw else {}
            if not isinstance(parsed_rules, dict):
                parsed_rules = {}
        except json.JSONDecodeError:
            parsed_rules = {}
        table_specs[object_id] = TableSpec(
            object_id=object_id,
            has_multilevel_headers=_as_bool(row.get("has_multilevel_headers", False)),
            format_rules={_as_text(k): _as_text(v) for k, v in parsed_rules.items()},
            row_label_col=_as_text(row.get("row_label_col", "")),
        )

    source_specs: dict[str, SourceSpec] = {}
    if not source_df.empty:
        for row in source_df.to_dict("records"):
            object_id = _as_text(row.get("object_id", ""))
            if object_id not in registry:
                continue
            source_specs[object_id] = SourceSpec(
                object_id=object_id,
                table_id=_as_text(row.get("table_id", "")),
                url=_as_text(row.get("url", "")),
                data_year=_as_int(row.get("data_year", 2023), default=2023),
                estimate_type=_as_text(row.get("estimate_type", "5-Year Estimates")),
                citation_month=_as_text(row.get("citation_month", "April")),
                citation_year=_as_int(row.get("citation_year", 2025), default=2025),
                custom_text=_as_text(row.get("custom_text", "")),
                source_text=_as_text(row.get("source_text", "")),
                note_text=_as_text(row.get("note_text", "")),
            )

    data_frames: dict[str, pd.DataFrame] = {}
    for record in registry.values():
        if record.data_sheet in sheets:
            data_frames[record.data_sheet] = sheets[record.data_sheet].copy()

    return WorkbookModel(
        workbook_path=source_path,
        registry=registry,
        figure_specs=figure_specs,
        table_specs=table_specs,
        source_specs=source_specs,
        data_frames=data_frames,
    )


def load_cha_workbook(workbook_path: str | Path) -> WorkbookModel:
    source_path = Path(workbook_path)
    sheets = _read_excel_raw(source_path)
    required = {"_registry", "_figure_specs", "_table_specs"}
    if required.issubset(sheets.keys()):
        return _load_normalized_workbook(source_path)
    return _load_flat_workbook(source_path)

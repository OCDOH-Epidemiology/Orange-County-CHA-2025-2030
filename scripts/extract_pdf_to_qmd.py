#!/usr/bin/env python3
"""
Extract CHA PDF narrative into Quarto chapter draft fragments.

Uses the PDF text layer (pdftotext) — not OCR — and maps Figure/Table numbers
to workbook Master ``F or T #`` → Object ID for include stubs.

Usage:
    # Printed page numbers (footer), recommended:
    python3 scripts/extract_pdf_to_qmd.py \\
      --pdf "/path/to/CHA.pdf" \\
      --from-page 196 --to-page 228 \\
      --out drafts/pdf-extract/ch08-remainder.qmd

    # Raw PDF page indices (1-based):
    python3 scripts/extract_pdf_to_qmd.py \\
      --pdf "/path/to/CHA.pdf" \\
      --from-page 202 --to-page 234 --pdf-pages \\
      --out drafts/pdf-extract/ch08-remainder.qmd

    # Append draft to an existing chapter after review:
    python3 scripts/extract_pdf_to_qmd.py ... --append chapters/08-leading-health-issues.qmd

Page numbering
--------------
This CHA PDF's printed footer page ≈ PDF page − 6 (e.g. printed 196 = PDF 202).
The script auto-calibrates the offset by scanning for a printed footer near
``--from-page`` unless ``--page-offset`` is given.

Include rules
-------------
- Figure N  → HTML comment + fig-<base>.qmd then tbl-<base>.qmd
- Table N  → HTML comment + tbl-<base>.qmd only
- Prose is reflowed (soft wraps / end-of-line hyphens only); wording unchanged
- Chart axes and table grids from the PDF are discarded (workbook renders them)
"""

from __future__ import annotations

import argparse
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKBOOK = PROJECT_ROOT / "data" / "raw" / "workbook.xlsx"
DEFAULT_OBJECTS_DIR = PROJECT_ROOT / "chapters" / "_generated" / "objects"
DEFAULT_INCLUDE_DIR = "_generated/objects"

# ALL-CAPS PDF titles that become ## (major sections under Leading Health Issues+)
MAJOR_SECTION_TITLES = {
    "CHRONIC DISEASE",
    "CANCER",
    "COMMUNICABLE DISEASE",
    "SEXUALLY TRANSMITTED INFECTIONS",
    "MATERNAL AND CHILD HEALTH",
    "HEALTH BEHAVIORS",
    "SUBSTANCE USE",
    "INJURY PREVENTION & SAFETY",
    "INJURY PREVENTION AND SAFETY",
    "PREVENTATIVE SCREENINGS",
    "PREVENTIVE SCREENINGS",
    "HANLON METHOD",
    "HEALTH SUMMIT",
    "COMMUNITY INPUT",
    "TRIANGULATE THE DATA",
    "COMMUNITY HEALTH IMPROVEMENT FOUNDATION",
    "APPENDICES",
    "SOCIAL DETERMINANTS OF HEALTH",
    "DEMOGRAPHICS",
    "MORTALITY",  # only when top-level; under MCH we special-case below
}

# When these appear under Maternal and Child Health, keep as ### not ##
MCH_SUBSECTIONS = {
    "FERTILITY",
    "TEEN BIRTHS",
    "MEDICAID",
    "MORTALITY",
    "PRENATAL CARE",
    "BIRTH OUTCOMES: PREMATURE BIRTH AND LOW BIRTHWEIGHT",
}

FIG_TABLE_START_RE = re.compile(r"^(Figure|Table)\s+(\d+)\s*$", re.IGNORECASE)
CITATION_RE = re.compile(
    r"\b(?:Figures?\s+F?|Tables?\s+T?)(\d+)(?:\s*[–—-]\s*(?:F|T)?(\d+))?",
    re.IGNORECASE,
)
PRINTED_FOOTER_RE = re.compile(r"^\d{1,3}$")
ALL_CAPS_TITLE_RE = re.compile(r"^[A-Z0-9][A-Z0-9 &/:’',\-]{2,}$")


@dataclass
class MasterObject:
    kind: str  # "F" or "T"
    number: int
    object_id: str
    caption: str


@dataclass
class PendingObject:
    kind: str  # "Figure" or "Table"
    number: int
    caption_hint: str = ""


@dataclass
class SectionBuffer:
    heading: str | None = None
    paragraphs: list[str] = field(default_factory=list)
    objects: list[PendingObject] = field(default_factory=list)


@dataclass
class ExtractReport:
    page_offset: int = 0
    pdf_from: int = 0
    pdf_to: int = 0
    unmatched: list[str] = field(default_factory=list)
    missing_files: list[str] = field(default_factory=list)
    emitted_objects: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Master sheet
# ---------------------------------------------------------------------------

def load_master_map(workbook_path: Path) -> dict[tuple[str, int], MasterObject]:
    """Map (F|T, number) → MasterObject from the workbook Master sheet."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Master" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"No 'Master' sheet in {workbook_path}")
    ws = wb["Master"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    mapping: dict[tuple[str, int], MasterObject] = {}
    for row in rows[1:]:
        ft = str(row[4] or "").strip()
        m = re.match(r"^([FT])(\d+)$", ft, re.IGNORECASE)
        if not m:
            continue
        kind = m.group(1).upper()
        number = int(m.group(2))
        object_id = str(row[6] or "").strip()
        caption = str(row[5] or "").strip()
        if not object_id:
            continue
        mapping[(kind, number)] = MasterObject(
            kind=kind, number=number, object_id=object_id, caption=caption
        )
    return mapping


def object_base_id(object_id: str) -> str:
    for prefix in ("tbl-", "fig-", "tlb-"):
        if object_id.lower().startswith(prefix):
            return object_id[len(prefix) :]
    return object_id


def _ascii_fold_id(value: str) -> str:
    """Fold Master IDs to on-disk object filenames (ï→empty→hyphen quirks)."""
    # NFKD then drop combining marks; also map common Master typos.
    import unicodedata

    folded = unicodedata.normalize("NFKD", value)
    folded = "".join(ch for ch in folded if not unicodedata.combining(ch))
    # Generator turned "naïve" into "na-ve" (ï removed leaving na + ve with hyphen)
    folded = folded.replace("naive", "na-ve")
    return folded


def resolve_object_filenames(
    object_id: str,
    objects_dir: Path,
    kind: str,
) -> tuple[str | None, str | None]:
    """
    Return (fig_filename, tbl_filename) that exist under objects_dir.

    Handles Master quirks: ``tbl-opioid-naïve-…``, ``tlb-past-30day-su``
    (on disk as ``fig-tlb-past-30day-su`` / ``tbl-tlb-past-30day-su``).
    """
    raw = object_id.strip()
    raw_folded = _ascii_fold_id(raw)
    base = object_base_id(raw)
    base_folded = _ascii_fold_id(base)

    fig_candidates: list[str] = []
    tbl_candidates: list[str] = []

    # Preserve typo prefix tlb- as part of the stem when files were generated that way
    if raw.lower().startswith("tlb-"):
        fig_candidates += [f"fig-{raw}.qmd", f"fig-{raw_folded}.qmd"]
        tbl_candidates += [f"tbl-{raw}.qmd", f"tbl-{raw_folded}.qmd", f"{raw}.qmd"]

    fig_candidates += [
        f"fig-{base_folded}.qmd",
        f"fig-{base}.qmd",
        f"fig-{raw_folded.removeprefix('tbl-').removeprefix('fig-')}.qmd",
    ]
    tbl_candidates += [
        f"tbl-{base_folded}.qmd",
        f"tbl-{base}.qmd",
        f"{raw_folded}.qmd" if raw_folded.startswith("tbl-") else f"tbl-{raw_folded}.qmd",
        f"{raw}.qmd",
    ]

    def first_existing(cands: list[str]) -> str | None:
        seen: set[str] = set()
        for name in cands:
            if name in seen:
                continue
            seen.add(name)
            if (objects_dir / name).exists():
                return name
        return None

    fig_name = first_existing(fig_candidates) if kind == "F" else None
    tbl_name = first_existing(tbl_candidates)
    if kind == "F" and fig_name is None:
        # still try tbl pairing stem from resolved tbl
        pass
    return fig_name, tbl_name


# ---------------------------------------------------------------------------
# PDF extraction + page offset
# ---------------------------------------------------------------------------

def require_pdftotext() -> str:
    path = shutil.which("pdftotext")
    if not path:
        raise SystemExit(
            "pdftotext not found. Install poppler (e.g. `brew install poppler`)."
        )
    return path


def pdf_page_count(pdf_path: Path, pdftotext: str) -> int:
    # pdfinfo is optional; fall back to binary search via pdftotext errors
    pdfinfo = shutil.which("pdfinfo")
    if pdfinfo:
        out = subprocess.check_output([pdfinfo, str(pdf_path)], text=True)
        for line in out.splitlines():
            if line.startswith("Pages:"):
                return int(line.split(":", 1)[1].strip())
    # crude fallback
    lo, hi = 1, 2000
    last_ok = 1
    while lo <= hi:
        mid = (lo + hi) // 2
        proc = subprocess.run(
            [pdftotext, "-f", str(mid), "-l", str(mid), str(pdf_path), "-"],
            capture_output=True,
            text=True,
        )
        if proc.returncode == 0:
            last_ok = mid
            lo = mid + 1
        else:
            hi = mid - 1
    return last_ok


def extract_pdf_page(pdf_path: Path, page: int, pdftotext: str) -> str:
    return subprocess.check_output(
        [pdftotext, "-f", str(page), "-l", str(page), "-layout", str(pdf_path), "-"],
        text=True,
    )


def calibrate_page_offset(
    pdf_path: Path,
    printed_from: int,
    pdftotext: str,
    n_pages: int,
    explicit_offset: int | None = None,
) -> int:
    """Return PDF_page - printed_page offset (typically 6 for this CHA)."""
    if explicit_offset is not None:
        return explicit_offset

    # Search near expected location for a page whose footer matches printed_from
    guess = printed_from + 6
    for pdf_page in range(max(1, guess - 15), min(n_pages, guess + 16) + 1):
        text = extract_pdf_page(pdf_path, pdf_page, pdftotext)
        footer = _page_footer(text)
        if footer == printed_from:
            return pdf_page - printed_from

    # Fallback: scan for SYPHILIS landmark (printed 196 in this CHA)
    for pdf_page in range(1, min(n_pages, 250) + 1):
        text = extract_pdf_page(pdf_path, pdf_page, pdftotext)
        first = _first_content_line(text)
        if first == "SYPHILIS":
            footer = _page_footer(text)
            if footer is not None:
                return pdf_page - footer
            return pdf_page - 196

    print(
        "Warning: could not calibrate page offset; defaulting to 6 "
        "(printed ≈ PDF − 6).",
        file=sys.stderr,
    )
    return 6


def _nonempty_lines(text: str) -> list[str]:
    return [ln.rstrip() for ln in text.splitlines() if ln.strip()]


def _first_content_line(text: str) -> str:
    lines = _nonempty_lines(text)
    return lines[0].strip() if lines else ""


def _page_footer(text: str) -> int | None:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return None
    # Prefer last short numeric line
    for ln in reversed(lines[-5:]):
        if PRINTED_FOOTER_RE.match(ln):
            return int(ln)
    return None


# ---------------------------------------------------------------------------
# Classification + narrative reflow
# ---------------------------------------------------------------------------

def title_case_heading(raw: str) -> str:
    """Convert ALL-CAPS PDF title to Title Case, preserving short words lightly."""
    # Keep colon segments
    parts = []
    for word in raw.strip().split():
        if word in {"&", "/", "–", "-"}:
            parts.append(word)
            continue
        # Preserve acronyms of length <= 4 that are all caps already handled via title()
        lowered = word.lower()
        if lowered in {"and", "or", "of", "the", "to", "in", "by", "for", "a", "an"}:
            parts.append(lowered if parts else lowered.capitalize())
        else:
            parts.append(word.capitalize())
    # Fix common patterns
    text = " ".join(parts)
    text = text.replace(" Hiv", " HIV").replace(" Aids", " AIDS")
    text = text.replace(" Zip", " ZIP").replace(" Nys", " NYS")
    text = text.replace(" OcdoH", " OCDOH").replace(" Prep", " PrEP")
    return text


def heading_markdown(title_raw: str, under_mch: bool) -> str:
    key = title_raw.strip().upper()
    if under_mch and key in MCH_SUBSECTIONS:
        return f"### {title_case_heading(title_raw)}"
    if key in MAJOR_SECTION_TITLES and not (under_mch and key == "MORTALITY"):
        return f"## {title_case_heading(title_raw)}"
    return f"### {title_case_heading(title_raw)}"


def is_all_caps_section_title(line: str) -> bool:
    s = line.strip()
    if len(s) < 3 or len(s) > 90:
        return False
    if not ALL_CAPS_TITLE_RE.match(s):
        return False
    # Reject lines that look like data / sources
    if s.startswith("SOURCE") or s.startswith("NOTE"):
        return False
    if re.search(r"\d{4}", s) and ":" not in s and "BIRTH" not in s:
        # years in titles are ok for birth outcomes; bare year lines no
        if re.fullmatch(r"\d{4}", s):
            return False
    return True


def reflow_paragraph_lines(lines: list[str]) -> str:
    """Join soft-wrapped lines.

    When a line ends with ``-`` and the next line continues with a letter, keep
    the hyphen and join with no space (``age-`` + ``adjusted`` → ``age-adjusted``,
    ``non-`` + ``Hispanic`` → ``non-Hispanic``). Soft mid-word breaks that used a
    hyphen are uncommon in this CHA PDF relative to real compound hyphens.
    """
    if not lines:
        return ""
    merged = lines[0].strip()
    for nxt_line in lines[1:]:
        nxt = nxt_line.strip()
        if merged.endswith("-") and nxt and nxt[0].isalpha():
            merged = merged + nxt
        else:
            merged = merged + " " + nxt
    return merged


def paragraph_groups_from_page_text(block: str) -> list[list[str]]:
    """Split a narrative block on blank lines into paragraph line-groups."""
    groups: list[list[str]] = []
    current: list[str] = []
    for ln in block.splitlines():
        if not ln.strip():
            if current:
                groups.append(current)
                current = []
            continue
        # Drop printed page footer if present as lone number
        if PRINTED_FOOTER_RE.match(ln.strip()) and len(current) > 3:
            continue
        current.append(ln.rstrip())
    if current:
        groups.append(current)
    return groups


# ---------------------------------------------------------------------------
# Page parsing
# ---------------------------------------------------------------------------

def strip_object_body_after(lines: list[str], start_idx: int) -> tuple[list[str], PendingObject | None]:
    """If lines[start_idx] is Figure/Table N, return (caption_hint, PendingObject)."""
    m = FIG_TABLE_START_RE.match(lines[start_idx].strip())
    if not m:
        return [], None
    kind = m.group(1).capitalize()
    number = int(m.group(2))
    # Collect following non-empty title lines until a pure number / axis-ish line
    title_parts: list[str] = []
    for ln in lines[start_idx + 1 : start_idx + 8]:
        s = ln.strip()
        if not s:
            if title_parts:
                break
            continue
        if FIG_TABLE_START_RE.match(s):
            break
        if re.fullmatch(r"\d+(\.\d+)?", s):
            break
        if s.lower().startswith(("source:", "note:", "s:")):
            break
        # Skip very short axis labels
        if s in {"Percent", "Rate per 100,000", "Number of Cases"}:
            break
        title_parts.append(s)
    hint = " ".join(title_parts)
    return title_parts, PendingObject(kind=kind, number=number, caption_hint=hint)


def parse_page(text: str) -> tuple[str, list[str], list[PendingObject]]:
    """
    Returns (page_kind, narrative_paragraph_texts, objects_on_page).

    page_kind: 'object' if page starts with Figure/Table, else 'narrative'
    """
    lines = _nonempty_lines(text)
    if not lines:
        return "empty", [], []

    first = lines[0].strip()
    objects: list[PendingObject] = []

    # Pure object page
    if FIG_TABLE_START_RE.match(first):
        _, obj = strip_object_body_after(lines, 0)
        if obj:
            objects.append(obj)
        # Multi-object rare; scan for additional Figure/Table headers
        for i, ln in enumerate(lines[1:], start=1):
            if FIG_TABLE_START_RE.match(ln.strip()):
                _, obj2 = strip_object_body_after(lines, i)
                if obj2:
                    objects.append(obj2)
        return "object", [], objects

    # Narrative page — may contain an embedded Table/Figure mid-page
    # Rebuild with blanks from original for paragraph detection
    # Cut narrative at first Figure/Table start line
    cut_idx = None
    raw_lines = text.splitlines()
    content_indices = [i for i, ln in enumerate(raw_lines) if ln.strip()]
    for pos, i in enumerate(content_indices):
        if FIG_TABLE_START_RE.match(raw_lines[i].strip()):
            # Don't cut if it's the section title somehow
            cut_idx = i
            _, obj = strip_object_body_after(
                [raw_lines[j] for j in content_indices], pos
            )
            if obj:
                objects.append(obj)
            # further objects after cut
            for pos2, j in enumerate(content_indices):
                if j <= i:
                    continue
                if FIG_TABLE_START_RE.match(raw_lines[j].strip()):
                    # map pos2 relative — easier scan nonempty after cut
                    pass
            break

    if cut_idx is not None:
        narrative_text = "\n".join(raw_lines[:cut_idx])
        # Collect all objects from cut onward
        nonempty_after = [ln for ln in raw_lines[cut_idx:] if ln.strip()]
        objects = []
        for i, ln in enumerate(nonempty_after):
            if FIG_TABLE_START_RE.match(ln.strip()):
                _, obj = strip_object_body_after(nonempty_after, i)
                if obj:
                    objects.append(obj)
    else:
        narrative_text = text

    # Drop trailing footer number
    paras: list[str] = []
    for group in paragraph_groups_from_page_text(narrative_text):
        # Skip lone footer groups
        if len(group) == 1 and PRINTED_FOOTER_RE.match(group[0].strip()):
            continue
        # If last line of group is footer, drop it
        if group and PRINTED_FOOTER_RE.match(group[-1].strip()):
            group = group[:-1]
        if not group:
            continue
        paras.append(reflow_paragraph_lines(group))

    return "narrative", paras, objects


def citation_order(prose: str) -> list[tuple[str, int]]:
    """Extract (F|T, number) mentions in prose order (ranges expanded)."""
    ordered: list[tuple[str, int]] = []
    seen: set[tuple[str, int]] = set()
    for m in CITATION_RE.finditer(prose):
        # Determine kind from the matched prefix text
        start = m.start()
        prefix = prose[max(0, start - 8) : m.start() + 10].lower()
        kind = "T" if "table" in prefix else "F"
        # More precise: look at the word before the number
        window = prose[max(0, start - 20) : m.end()].lower()
        if re.search(r"tables?\s+t?\d", window):
            kind = "T"
        elif re.search(r"figures?\s+f?\d", window):
            kind = "F"
        a = int(m.group(1))
        b = int(m.group(2)) if m.group(2) else a
        # If range spans figures cited as Figures 110–116, kind is F
        for n in range(min(a, b), max(a, b) + 1):
            key = (kind, n)
            if key not in seen:
                seen.add(key)
                ordered.append(key)
    return ordered


# ---------------------------------------------------------------------------
# Emit Quarto
# ---------------------------------------------------------------------------

def render_includes(
    objects: list[PendingObject],
    master: dict[tuple[str, int], MasterObject],
    objects_dir: Path,
    include_dir: str,
    report: ExtractReport,
    prefer_order: list[tuple[str, int]] | None = None,
) -> list[str]:
    """Emit HTML comments + includes for pending objects."""
    # Deduplicate preserving order
    pending_keys: list[tuple[str, int]] = []
    hints: dict[tuple[str, int], str] = {}
    for obj in objects:
        key = ("F" if obj.kind.lower().startswith("fig") else "T", obj.number)
        hints[key] = obj.caption_hint
        if key not in pending_keys:
            pending_keys.append(key)

    # Match chapter style: tables first (ascending), then figures (ascending).
    # prefer_order from prose is used only to break ties within the same kind.
    prefer_index = {key: i for i, key in enumerate(prefer_order or [])}

    def sort_key(key: tuple[str, int]) -> tuple[int, int, int]:
        kind, number = key
        kind_rank = 0 if kind == "T" else 1
        prose_rank = prefer_index.get(key, 10_000)
        return (kind_rank, number, prose_rank)

    pending_keys = sorted(pending_keys, key=sort_key)

    lines: list[str] = []
    for kind, number in pending_keys:
        rec = master.get((kind, number))
        label = "Figure" if kind == "F" else "Table"
        if rec is None:
            report.unmatched.append(f"{label} {number}")
            caption = hints.get((kind, number), "")
            lines.append(f"<!-- {label} {number}: {caption} [UNMATCHED IN MASTER] -->")
            lines.append("")
            continue

        base = object_base_id(rec.object_id)
        caption = rec.caption or hints.get((kind, number), "")
        lines.append(f"<!-- {label} {number}: {caption} -->")

        fig_name, tbl_name = resolve_object_filenames(rec.object_id, objects_dir, kind)

        if kind == "F":
            if fig_name is None:
                report.missing_files.append(f"fig-{_ascii_fold_id(base)}.qmd")
                lines.append(f"<!-- MISSING object file for Figure {number} ({rec.object_id}) -->")
            else:
                lines.append(f"{{{{< include {include_dir}/{fig_name} >}}}}")
            if tbl_name is None:
                report.missing_files.append(f"tbl-{_ascii_fold_id(base)}.qmd")
                lines.append(f"<!-- MISSING paired table file for Figure {number} ({rec.object_id}) -->")
            else:
                lines.append(f"{{{{< include {include_dir}/{tbl_name} >}}}}")
            report.emitted_objects.append(f"F{number}:{base}")
        else:
            if tbl_name is None:
                report.missing_files.append(f"tbl-{_ascii_fold_id(base)}.qmd")
                lines.append(f"<!-- MISSING object file for Table {number} ({rec.object_id}) -->")
            else:
                lines.append(f"{{{{< include {include_dir}/{tbl_name} >}}}}")
            report.emitted_objects.append(f"T{number}:{base}")
        lines.append("")
    return lines


def flush_section(
    buf: SectionBuffer,
    master: dict[tuple[str, int], MasterObject],
    objects_dir: Path,
    include_dir: str,
    report: ExtractReport,
    out_lines: list[str],
) -> None:
    if not buf.heading and not buf.paragraphs and not buf.objects:
        return

    if buf.heading:
        out_lines.append(buf.heading)
        out_lines.append("")

    prose_blocks = [p for p in buf.paragraphs if p.strip()]
    # Drop paragraph if it is only the ALL-CAPS title duplicated
    for p in prose_blocks:
        out_lines.append(p.strip())
        out_lines.append("")

    prose_join = " ".join(prose_blocks)
    prefer = citation_order(prose_join) if prose_join else None
    include_lines = render_includes(
        buf.objects, master, objects_dir, include_dir, report, prefer_order=prefer
    )
    out_lines.extend(include_lines)


def extract_range(
    pdf_path: Path,
    pdf_from: int,
    pdf_to: int,
    master: dict[tuple[str, int], MasterObject],
    objects_dir: Path,
    include_dir: str,
    pdftotext: str,
    stop_at_titles: set[str] | None = None,
) -> tuple[str, ExtractReport]:
    stop_at_titles = {t.upper() for t in (stop_at_titles or set())}
    report = ExtractReport(pdf_from=pdf_from, pdf_to=pdf_to)
    out_lines: list[str] = []
    buf = SectionBuffer()
    under_mch = False
    # Do not stop on the first title in-range (e.g. extracting Health Behaviors
    # itself). Only stop when a later major-section title matches.
    seen_section_title = False

    for page in range(pdf_from, pdf_to + 1):
        text = extract_pdf_page(pdf_path, page, pdftotext)
        kind, paras, objects = parse_page(text)

        if kind == "empty":
            continue

        if kind == "object":
            # Attach objects to current section buffer
            buf.objects.extend(objects)
            continue

        # narrative
        i = 0
        while i < len(paras):
            p = paras[i].strip()
            if not p:
                i += 1
                continue

            # Section title detection: first paragraph is a single ALL-CAPS line
            # or starts with ALL-CAPS title line
            first_line = p.split("\n")[0].strip() if "\n" in p else p
            # After reflow, paragraph is one line. Check if whole para is a title.
            is_title = is_all_caps_section_title(p) or (
                is_all_caps_section_title(first_line) and len(p) < 100
            )

            if is_title:
                title_key = p.strip().upper()
                if title_key in stop_at_titles and seen_section_title:
                    flush_section(buf, master, objects_dir, include_dir, report, out_lines)
                    buf = SectionBuffer()
                    return "\n".join(out_lines).rstrip() + "\n", report

                # New section — flush previous
                flush_section(buf, master, objects_dir, include_dir, report, out_lines)
                seen_section_title = True
                if title_key == "MATERNAL AND CHILD HEALTH":
                    under_mch = True
                # Chapter YAML already has the top-level title; skip duplicating it.
                if title_key in {
                    "HEALTH BEHAVIORS",
                    "HANLON METHOD",
                    "HEALTH SUMMIT",
                    "COMMUNITY INPUT",
                }:
                    # Keep intro paragraphs under this title without a ## heading.
                    buf = SectionBuffer(heading=None)
                    i += 1
                    continue
                heading = heading_markdown(p.strip(), under_mch=under_mch)
                buf = SectionBuffer(heading=heading)
                i += 1
                continue

            # If paragraph begins with an ALL-CAPS title followed by prose
            # (rare after reflow), try split
            words = p.split()
            if words and words[0].isupper() and len(words[0]) > 3:
                # e.g. shouldn't happen after good reflow of multi-line titles
                pass

            buf.paragraphs.append(p)
            i += 1

        if objects:
            buf.objects.extend(objects)

    flush_section(buf, master, objects_dir, include_dir, report, out_lines)
    return "\n".join(out_lines).rstrip() + "\n", report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Extract CHA PDF narrative to Quarto draft with figure/table includes."
    )
    p.add_argument("--pdf", type=Path, required=True, help="Path to CHA PDF")
    p.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="CHA workbook with Master sheet (default: data/raw/workbook.xlsx)",
    )
    p.add_argument(
        "--from-page",
        type=int,
        required=True,
        help="Start page (printed footer number unless --pdf-pages)",
    )
    p.add_argument(
        "--to-page",
        type=int,
        required=True,
        help="End page inclusive (printed footer number unless --pdf-pages)",
    )
    p.add_argument(
        "--pdf-pages",
        action="store_true",
        help="Interpret --from-page/--to-page as 1-based PDF page indices",
    )
    p.add_argument(
        "--page-offset",
        type=int,
        default=None,
        help="PDF_page − printed_page (auto-calibrated if omitted; typically 6)",
    )
    p.add_argument(
        "--out",
        type=Path,
        required=True,
        help="Output draft .qmd path",
    )
    p.add_argument(
        "--append",
        type=Path,
        default=None,
        help="If set, append draft content to this chapter file after writing --out",
    )
    p.add_argument(
        "--objects-dir",
        type=Path,
        default=DEFAULT_OBJECTS_DIR,
        help="Directory of generated object .qmd files",
    )
    p.add_argument(
        "--include-dir",
        type=str,
        default=DEFAULT_INCLUDE_DIR,
        help="Include path prefix used inside {{< include >}}",
    )
    p.add_argument(
        "--report",
        type=Path,
        default=None,
        help="Optional path to write a text extraction report",
    )
    p.add_argument(
        "--stop-at",
        action="append",
        default=[],
        help=(
            "ALL-CAPS section title that ends extraction when encountered after "
            "content has started (repeatable). Example: --stop-at 'HANLON METHOD'"
        ),
    )
    return p


def format_report(report: ExtractReport, page_offset: int) -> str:
    lines = [
        "PDF → Quarto extraction report",
        f"  page_offset (PDF − printed): {page_offset}",
        f"  PDF pages: {report.pdf_from}–{report.pdf_to}",
        f"  objects emitted: {len(report.emitted_objects)}",
    ]
    if report.emitted_objects:
        lines.append("  " + ", ".join(report.emitted_objects[:40]))
        if len(report.emitted_objects) > 40:
            lines.append(f"  … and {len(report.emitted_objects) - 40} more")
    if report.unmatched:
        lines.append("  UNMATCHED in Master:")
        for u in report.unmatched:
            lines.append(f"    - {u}")
    else:
        lines.append("  unmatched: none")
    if report.missing_files:
        lines.append("  MISSING object files:")
        for f in sorted(set(report.missing_files)):
            lines.append(f"    - {f}")
    else:
        lines.append("  missing object files: none")
    return "\n".join(lines) + "\n"


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    pdf_path: Path = args.pdf.expanduser().resolve()
    if not pdf_path.exists():
        raise SystemExit(f"PDF not found: {pdf_path}")
    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    pdftotext = require_pdftotext()
    n_pages = pdf_page_count(pdf_path, pdftotext)

    if args.pdf_pages:
        pdf_from, pdf_to = args.from_page, args.to_page
        page_offset = 0
    else:
        page_offset = calibrate_page_offset(
            pdf_path, args.from_page, pdftotext, n_pages, args.page_offset
        )
        pdf_from = args.from_page + page_offset
        pdf_to = args.to_page + page_offset

    if pdf_from < 1 or pdf_to > n_pages or pdf_from > pdf_to:
        raise SystemExit(
            f"Invalid page range PDF {pdf_from}–{pdf_to} (document has {n_pages} pages; "
            f"offset={page_offset})"
        )

    master = load_master_map(args.workbook)
    body, report = extract_range(
        pdf_path=pdf_path,
        pdf_from=pdf_from,
        pdf_to=pdf_to,
        master=master,
        objects_dir=args.objects_dir,
        include_dir=args.include_dir.rstrip("/"),
        pdftotext=pdftotext,
        stop_at_titles=set(args.stop_at or []),
    )
    report.page_offset = page_offset

    args.out.parent.mkdir(parents=True, exist_ok=True)
    header = (
        f"<!-- Generated by scripts/extract_pdf_to_qmd.py\n"
        f"     PDF: {pdf_path.name}\n"
        f"     printed pages: {args.from_page}–{args.to_page}"
        f"{'' if args.pdf_pages else f' (offset {page_offset} → PDF {pdf_from}–{pdf_to})'}\n"
        f"     Review before merging into a chapter. -->\n\n"
    )
    args.out.write_text(header + body, encoding="utf-8")
    print(f"Wrote draft: {args.out}")

    report_text = format_report(report, page_offset)
    print(report_text, end="")
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(report_text, encoding="utf-8")
        print(f"Wrote report: {args.report}")

    if args.append:
        chapter = args.append
        existing = chapter.read_text(encoding="utf-8")
        # Ensure separation
        sep = "\n" if existing.endswith("\n") else "\n\n"
        chapter.write_text(existing + sep + body, encoding="utf-8")
        print(f"Appended draft body to: {chapter}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
